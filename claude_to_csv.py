import requests
from bs4 import BeautifulSoup
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Lista URL-i do testów
test_urls = [
    "https://audiblask.olx.pl/home/",
    "https://media-max.olx.pl/home/",
    "https://sprzedazalufelg.olx.pl/home/",
    "https://www.olx.pl/oferty/uzytkownik/42D2P/",
    "https://brickmarket.olx.pl/home/",
    "https://autoczesci.olx.pl/home/",
    "https://skotniki.olx.pl/home/",
    "https://www.olx.pl/oferty/uzytkownik/2UafH/",
    "https://motozbyt.olx.pl/home/",
    "https://www.olx.pl/oferty/uzytkownik/Yxbr/",
    "https://www.olx.pl/oferty/uzytkownik/1DsEB/",
    "https://www.olx.pl/oferty/uzytkownik/1qlFD/"
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "pl-PL,pl;q=0.9"
}


def get_shop_name_from_url(shop_url):
    """
    Pobiera nazwę sklepu/użytkownika ze strony
    """
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')

    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get(shop_url)
        time.sleep(3)

        # Dla sklepów firmowych
        if ".olx.pl/home/" in shop_url:
            # Szukaj w tytule strony
            title = driver.title
            if " - " in title:
                name = title.split(" - ")[0].strip()
                if name and len(name) < 100 and "OLX" not in name:
                    return name

            # Szukaj w h1
            try:
                h1 = driver.find_element(By.TAG_NAME, "h1")
                if h1 and h1.text.strip():
                    return h1.text.strip()
            except:
                pass

        # Dla użytkowników - ulepszone szukanie
        elif "/oferty/uzytkownik/" in shop_url:
            # Lista fraz do pominięcia (elementy interfejsu)
            skip_phrases = [
                'Czat', 'Powiadomienia', 'Twoje konto', 'Dodaj ogłoszenie',
                'Obserwuj', 'Podziel się', 'Dbamy o Twoją prywatność',
                'Ogłoszenia', 'Oceny', 'Filtruj ogłoszenia', 'Znajdź na tej stronie',
                'Sortuj', 'Najnowsze', 'Kategorie', 'Wszystkie ogłoszenia',
                'Znaleźliśmy', 'ogłoszeń', 'Na OLX od', 'Ostatnio online'
            ]

            # Metoda 1: Szukaj nazwy między elementami nawigacji a oceną
            page_text = driver.find_element(By.TAG_NAME, "body").text
            lines = page_text.split('\n')

            # Znajdź indeks gdzie kończy się nawigacja (po "Dodaj ogłoszenie")
            start_index = 0
            for i, line in enumerate(lines):
                if "Dodaj ogłoszenie" in line:
                    start_index = i + 1
                    break

            # Szukaj nazwy od tego miejsca
            for i in range(start_index, min(start_index + 10, len(lines))):
                line = lines[i].strip()
                # Pomijamy puste linie i znane frazy
                if line and len(line) > 2 and len(line) < 50:
                    skip = False
                    for phrase in skip_phrases:
                        if phrase.lower() in line.lower():
                            skip = True
                            break

                    if not skip and not line.replace('.', '').replace('/', '').replace(' ', '').isdigit():
                        # Sprawdź czy następna linia to ocena
                        if i + 1 < len(lines) and ('/ 5' in lines[i + 1] or 'ocen' in lines[i + 1]):
                            print(f"    Debug: Znaleziono nazwę przed oceną: '{line}'")
                            return line

            # Metoda 2: Użyj JavaScript do precyzyjnego szukania
            try:
                name = driver.execute_script("""
                    // Znajdź sekcję z informacjami o użytkowniku
                    const mainContent = document.querySelector('main, [role="main"], .content');
                    if (!mainContent) return null;

                    // Szukaj elementów które mogą zawierać nazwę
                    const possibleElements = mainContent.querySelectorAll('h1, h2, h3, strong, b');

                    for (let el of possibleElements) {
                        const text = el.textContent.trim();

                        // Lista fraz do pominięcia
                        const skipPhrases = ['Filtruj', 'Ogłoszenia', 'Oceny', 'Sortuj', 
                                           'Kategorie', 'Znaleźliśmy', 'Wszystkie'];

                        // Sprawdź czy to może być nazwa
                        if (text && text.length > 2 && text.length < 50) {
                            let skip = false;
                            for (let phrase of skipPhrases) {
                                if (text.includes(phrase)) {
                                    skip = true;
                                    break;
                                }
                            }

                            if (!skip) {
                                // Sprawdź czy element jest widoczny i czy nie jest linkiem
                                const rect = el.getBoundingClientRect();
                                if (rect.width > 0 && rect.height > 0 && 
                                    el.tagName !== 'A' && !el.closest('a')) {

                                    // Sprawdź czy w pobliżu jest ocena (to wskazuje na nazwę użytkownika)
                                    const parent = el.parentElement;
                                    const parentText = parent ? parent.textContent : '';
                                    if (parentText.includes('/ 5') || parentText.includes('ocen')) {
                                        return text;
                                    }
                                }
                            }
                        }
                    }

                    return null;
                """)

                if name:
                    return name
            except:
                pass

            # Metoda 3: Szukaj w konkretnym obszarze strony
            try:
                # Czasami nazwa jest w sekcji przed przyciskami "Podziel się" i "Obserwuj"
                share_button = driver.find_element(By.XPATH, "//*[contains(text(), 'Podziel się')]")
                # Cofnij się do rodzica i szukaj nazwy
                parent = share_button.find_element(By.XPATH, "../..")
                texts = parent.find_elements(By.XPATH, ".//h1 | .//h2 | .//h3 | .//strong")

                for elem in texts:
                    text = elem.text.strip()
                    if text and len(text) > 2 and len(text) < 50:
                        skip = False
                        for phrase in skip_phrases:
                            if phrase.lower() in text.lower():
                                skip = True
                                break
                        if not skip:
                            return text
            except:
                pass

            # Ostatnia deska ratunku - wyciągnij z URL
            match = re.search(r'/uzytkownik/([^/]+)/', shop_url)
            if match:
                user_id = match.group(1)
                # Spróbuj pobrać z tytułu strony
                title = driver.title
                if " - " in title and "OLX" in title:
                    potential_name = title.split(" - ")[0].strip()
                    if potential_name and "ogłoszenia" not in potential_name.lower():
                        return potential_name

                return f"Użytkownik_{user_id}"

        # Fallback - nazwa z domeny
        if ".olx.pl" in shop_url:
            match = re.search(r'https://([^.]+)\.olx\.pl', shop_url)
            if match:
                return match.group(1).capitalize()

        return "Nieznany"

    except Exception as e:
        print(f"    Błąd pobierania nazwy: {str(e)[:100]}")
        return "Błąd"
    finally:
        driver.quit()


def ctc_get_olx_ads_count_selenium(shop_url):
    """
    Używa Selenium do pobrania liczby ogłoszeń
    """
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')

    driver = webdriver.Chrome(options=chrome_options)

    try:
        driver.get(shop_url)
        time.sleep(3)

        page_text = driver.find_element(By.TAG_NAME, "body").text

        # Dla użytkowników
        if "/oferty/uzytkownik/" in shop_url:
            # Sprawdź czy to nie jest przekierowanie do wszystkich ogłoszeń
            if "wszystkie ogłoszenia w" in page_text.lower():
                return 0

            match = re.search(r'Znaleźliśmy (\d+) ogłoszeń', page_text)
            if match:
                count = int(match.group(1))
                if count > 1000000:
                    return 0
                return count

            match = re.search(r'Wszystkie ogłoszenia\s*(\d+)', page_text)
            if match:
                count = int(match.group(1))
                if count > 1000000:
                    return 0
                return count

            if any(text in page_text for text in ["Brak ogłoszeń", "Nie ma ogłoszeń", "0 ogłoszeń"]):
                return 0

        driver.quit()
        return None

    except Exception as e:
        driver.quit()
        return None


def ctc_get_olx_ads_count(shop_url):
    """
    Pobiera liczbę ogłoszeń
    """
    resp = requests.get(shop_url, headers=HEADERS)
    if resp.status_code != 200:
        return None

    html = resp.text
    soup = BeautifulSoup(html, 'html.parser')

    is_user_page = "/oferty/uzytkownik/" in shop_url
    is_shop_page = ".olx.pl/home/" in shop_url

    if is_shop_page:
        # Dla sklepów firmowych
        for element in soup.find_all(['button', 'a', 'span', 'div', 'h1', 'h2', 'h3']):
            text = element.get_text().strip()

            patterns = [
                r'(\d+[\s\d]*)\s*ogłoszeń',
                r'(\d+[\s\d]*)\s*ofert',
            ]

            for pattern in patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    count = int(match.group(1).replace(' ', '').replace('\xa0', ''))
                    if 0 < count <= 10000:
                        return count

    elif is_user_page:
        return ctc_get_olx_ads_count_selenium(shop_url)

    if is_shop_page:
        return ctc_get_olx_ads_count_selenium(shop_url)

    return None

#Zmiana: results w argumencie zamiast urls
def process_urls_to_xlsx(urls, output_filename="olx_sellers.xlsx"):
    """
    Przetwarza listę URL-i i zapisuje wyniki do XLSX
    """
    results = []

    print(f"Przetwarzanie {len(urls)} URL-i...\n")

    for i, url in enumerate(urls, 1):
        print(f"\n[{i}/{len(urls)}] Przetwarzanie: {url}")

        # Pobierz nazwę
        print("  → Pobieram nazwę...")
        name = get_shop_name_from_url(url)
        print(f"  ✓ Nazwa: {name}")

        # Pobierz liczbę ogłoszeń
        print("  → Pobieram liczbę ogłoszeń...")
        ads_count = ctc_get_olx_ads_count(url)
        if ads_count is not None:
            print(f"  ✓ Liczba ogłoszeń: {ads_count}")
        else:
            ads_count = 0
            print(f"  ✗ Nie udało się pobrać liczby ogłoszeń")

        # Dodaj do wyników
        results.append({
            'Nazwa użytkownika/firmy': name,
            'Link do konta': url,
            'Nr telefonu': '',
            'Login': '',
            'Hasło': '',
            'Ilość ogłoszeń': ads_count,
            'Nazwa platformy': 'olx.pl'
        })

    # Utwórz DataFrame
    df = pd.DataFrame(results)

    # Zapisz do XLSX z formatowaniem
    print(f"\n\nZapisywanie wyników do {output_filename}...")

    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sprzedawcy OLX', index=False)

        # Pobierz arkusz
        worksheet = writer.sheets['Sprzedawcy OLX']

        # Formatowanie nagłówków
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Dostosuj szerokość kolumn
        column_widths = {
            'A': 30,  # Nazwa
            'B': 50,  # Link
            'C': 15,  # Nr telefonu
            'D': 15,  # Login
            'E': 15,  # Hasło
            'F': 15,  # Ilość ogłoszeń
            'G': 15  # Platforma
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        # Wyrównaj dane
        for row in worksheet.iter_rows(min_row=2):
            row[5].alignment = Alignment(horizontal="center")  # Ilość ogłoszeń
            row[6].alignment = Alignment(horizontal="center")  # Platforma

    print(f"✅ Zapisano {len(results)} rekordów do {output_filename}")

    # Podsumowanie
    print("\nPodsumowanie:")
    total_ads = df['Ilość ogłoszeń'].sum()
    valid_counts = df[df['Ilość ogłoszeń'] > 0]
    print(f"  - Łączna liczba ogłoszeń: {total_ads}")
    print(f"  - Średnia liczba ogłoszeń: {df['Ilość ogłoszeń'].mean():.1f}")
    print(f"  - Rekordy z ogłoszeniami: {len(valid_counts)}/{len(df)}")



if __name__ == '__main__':
    # Upewnij się, że masz zainstalowane wymagane pakiety
    try:
        import pandas
        import openpyxl
    except ImportError:
        print("Instaluję wymagane pakiety...")
        import subprocess

        subprocess.check_call(["pip", "install", "pandas", "openpyxl"])

    # Generuj nazwę pliku z datą i czasem
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"olx_sellers_{timestamp}.xlsx"

    # Przetwórz URL-e i zapisz do XLSX
    process_urls_to_xlsx(test_urls, output_file)